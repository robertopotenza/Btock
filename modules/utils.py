"""
Utility Functions for Btock Stock KPI Scoring Dashboard
Helper functions for file processing, data validation, and export
"""

import pandas as pd
import streamlit as st
import numpy as np
import io
from typing import List, Dict, Optional, Tuple
import re
import uuid
from datetime import datetime


class FileProcessor:
    """Handles file upload and processing operations"""
    
    @staticmethod
    def process_uploaded_file(uploaded_file) -> Optional[List[str]]:
        """
        Process uploaded Excel or CSV file to extract ticker list
        
        Args:
            uploaded_file: Streamlit uploaded file object
            
        Returns:
            List of ticker symbols or None if error
        """
        try:
            if uploaded_file is None:
                return None
            
            # Determine file type and read accordingly
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(uploaded_file)
            else:
                st.error("Unsupported file format. Please upload CSV or Excel files.")
                return None
            
            # Look for ticker column (case insensitive)
            ticker_column = None
            for col in df.columns:
                if col.lower() in ['ticker', 'symbol', 'stock', 'tickers', 'symbols']:
                    ticker_column = col
                    break
            
            if ticker_column is None:
                st.error("No 'Ticker' column found in the uploaded file. Please ensure your file has a column named 'Ticker'.")
                return None
            
            # Extract tickers and clean them
            tickers = df[ticker_column].dropna().astype(str).tolist()
            cleaned_tickers = []
            
            for ticker in tickers:
                # Clean ticker symbol
                cleaned_ticker = FileProcessor.clean_ticker_symbol(ticker)
                if cleaned_ticker:
                    cleaned_tickers.append(cleaned_ticker)
            
            if not cleaned_tickers:
                st.error("No valid ticker symbols found in the uploaded file.")
                return None
            
            # Remove duplicates while preserving order
            unique_tickers = list(dict.fromkeys(cleaned_tickers))
            
            st.success(f"Successfully loaded {len(unique_tickers)} unique ticker symbols.")
            
            return unique_tickers
            
        except Exception as e:
            st.error(f"Error processing uploaded file: {str(e)}")
            return None
    
    @staticmethod
    def clean_ticker_symbol(ticker: str) -> Optional[str]:
        """
        Clean and validate ticker symbol
        
        Args:
            ticker: Raw ticker symbol
            
        Returns:
            Cleaned ticker symbol or None if invalid
        """
        if not ticker or pd.isna(ticker):
            return None
        
        # Convert to string and strip whitespace
        ticker = str(ticker).strip().upper()
        
        # Remove common prefixes/suffixes and special characters
        ticker = re.sub(r'[^A-Z0-9.-]', '', ticker)
        
        # Basic validation - ticker should be 1-5 characters for most exchanges
        if len(ticker) < 1 or len(ticker) > 10:
            return None
        
        # Skip obviously invalid entries
        invalid_patterns = ['N/A', 'NULL', 'NONE', 'ERROR', '']
        if ticker in invalid_patterns:
            return None
        
        return ticker
    
    @staticmethod
    def create_results_excel(results_df: pd.DataFrame) -> bytes:
        """
        Create Excel file from results DataFrame
        
        Args:
            results_df: DataFrame with analysis results
            
        Returns:
            Excel file as bytes
        """
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write main results
                results_df.to_excel(writer, sheet_name='Stock Analysis Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Stock Analysis Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add formatting for signal column
                from openpyxl.styles import PatternFill
                
                # Find signal column
                signal_col = None
                for idx, col in enumerate(results_df.columns, 1):
                    if col.lower() == 'signal':
                        signal_col = idx
                        break
                
                if signal_col:
                    # Color code signals
                    buy_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
                    sell_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink
                    hold_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light yellow
                    
                    for row in range(2, len(results_df) + 2):  # Skip header row
                        cell = worksheet.cell(row=row, column=signal_col)
                        if cell.value == 'BUY':
                            cell.fill = buy_fill
                        elif cell.value == 'SELL':
                            cell.fill = sell_fill
                        elif cell.value == 'HOLD':
                            cell.fill = hold_fill
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error creating Excel file: {str(e)}")
            return b''


class WeightValidator:
    """Handles weight validation and normalization"""
    
    @staticmethod
    def validate_and_normalize_weights(weights: Dict[str, float]) -> Tuple[Dict[str, float], bool]:
        """
        Validate and normalize weights to sum to 1.0
        
        Args:
            weights: Dictionary of category weights
            
        Returns:
            Tuple of (normalized_weights, is_valid)
        """
        try:
            # Check if all weights are non-negative
            for category, weight in weights.items():
                if weight < 0:
                    st.error(f"Weight for {category} cannot be negative.")
                    return weights, False
            
            # Calculate total weight
            total_weight = sum(weights.values())
            
            if total_weight == 0:
                st.error("Total weight cannot be zero. Please assign positive weights.")
                return weights, False
            
            # Normalize weights to sum to 1.0
            normalized_weights = {k: v / total_weight for k, v in weights.items()}
            
            # Show normalization message if needed
            if abs(total_weight - 1.0) > 0.001:
                st.info(f"Weights normalized from total {total_weight:.3f} to 1.000")
            
            return normalized_weights, True
            
        except Exception as e:
            st.error(f"Error validating weights: {str(e)}")
            return weights, False
    
    @staticmethod
    def get_default_weights() -> Dict[str, float]:
        """Get default category weights"""
        return {
            'momentum': 0.30,
            'trend': 0.40,
            'volatility': 0.05,
            'strength': 0.05,
            'support_resistance': 0.20
        }


class SessionManager:
    """Manages analysis sessions and results"""
    
    @staticmethod
    def generate_session_id() -> str:
        """Generate unique session ID"""
        return str(uuid.uuid4())[:8]
    
    @staticmethod
    def create_session_summary(session_id: str, weights: Dict[str, float], 
                             total_tickers: int, successful_tickers: int) -> Dict:
        """
        Create session summary
        
        Args:
            session_id: Unique session identifier
            weights: Category weights used
            total_tickers: Total number of tickers processed
            successful_tickers: Number of successfully analyzed tickers
            
        Returns:
            Session summary dictionary
        """
        return {
            'session_id': session_id,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'weights': weights,
            'total_tickers': total_tickers,
            'successful_tickers': successful_tickers,
            'success_rate': (successful_tickers / total_tickers * 100) if total_tickers > 0 else 0
        }


class DataFormatter:
    """Handles data formatting and display"""
    
    @staticmethod
    def format_results_for_display(results: List[Dict]) -> pd.DataFrame:
        """
        Format analysis results for display in Streamlit
        
        Args:
            results: List of analysis result dictionaries
            
        Returns:
            Formatted DataFrame
        """
        try:
            if not results:
                return pd.DataFrame()
            
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Reorder columns for better display
            column_order = [
                'ticker', 'current_price', 'final_weighted_score', 'signal',
                'momentum_score', 'trend_score', 'volatility_score', 
                'strength_score', 'support_resistance_score', 'error_message'
            ]
            
            # Only include columns that exist
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]
            
            # Format numeric columns
            numeric_columns = [
                'current_price', 'final_weighted_score', 'momentum_score',
                'trend_score', 'volatility_score', 'strength_score', 'support_resistance_score'
            ]
            
            for col in numeric_columns:
                if col in df.columns:
                    if col == 'current_price':
                        df[col] = df[col].apply(lambda x: f"${x:.2f}" if pd.notna(x) else "N/A")
                    else:
                        df[col] = df[col].apply(lambda x: f"{x:.4f}" if pd.notna(x) else "N/A")
            
            # Rename columns for display
            column_names = {
                'ticker': 'Ticker',
                'current_price': 'Current Price',
                'final_weighted_score': 'Final Score',
                'signal': 'Signal',
                'momentum_score': 'Momentum',
                'trend_score': 'Trend',
                'volatility_score': 'Volatility',
                'strength_score': 'Strength',
                'support_resistance_score': 'Support/Resistance',
                'error_message': 'Error'
            }
            
            df = df.rename(columns=column_names)
            
            # Sort by Final Score (descending)
            if 'Final Score' in df.columns:
                # Convert back to numeric for sorting
                df['Final Score Numeric'] = df['Final Score'].str.replace('N/A', '0').astype(float)
                df = df.sort_values('Final Score Numeric', ascending=False)
                df = df.drop('Final Score Numeric', axis=1)
            
            return df
            
        except Exception as e:
            st.error(f"Error formatting results: {str(e)}")
            return pd.DataFrame()
    
    @staticmethod
    def create_summary_stats(results: List[Dict]) -> Dict:
        """
        Create summary statistics from results
        
        Args:
            results: List of analysis result dictionaries
            
        Returns:
            Summary statistics dictionary
        """
        try:
            if not results:
                return {}
            
            # Filter out results with errors
            valid_results = [r for r in results if 'error_message' not in r or not r['error_message']]
            
            if not valid_results:
                return {'total_analyzed': len(results), 'successful': 0, 'errors': len(results)}
            
            # Calculate signal distribution
            signals = [r['signal'] for r in valid_results]
            signal_counts = {
                'BUY': signals.count('BUY'),
                'HOLD': signals.count('HOLD'),
                'SELL': signals.count('SELL')
            }
            
            # Calculate score statistics
            scores = [r['final_weighted_score'] for r in valid_results]
            
            summary = {
                'total_analyzed': len(results),
                'successful': len(valid_results),
                'errors': len(results) - len(valid_results),
                'signal_distribution': signal_counts,
                'score_stats': {
                    'mean': np.mean(scores),
                    'median': np.median(scores),
                    'std': np.std(scores),
                    'min': np.min(scores),
                    'max': np.max(scores)
                }
            }
            
            return summary
            
        except Exception as e:
            st.error(f"Error creating summary stats: {str(e)}")
            return {}



class ExcelExporter:
    """Handles Excel file export with formatting"""
    
    @staticmethod
    def export_results(df: pd.DataFrame, filename: str = "btock_analysis_results.xlsx") -> bytes:
        """
        Export results to Excel with formatting
        
        Args:
            df: Results DataFrame
            filename: Output filename
            
        Returns:
            Excel file as bytes
        """
        try:
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Analysis Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Analysis Results']
                
                # Add conditional formatting for signals
                from openpyxl.styles import PatternFill
                from openpyxl.formatting.rule import CellIsRule
                
                # Define colors for signals
                buy_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
                hold_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light yellow
                sell_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light red
                
                # Find the Signal column
                signal_col = None
                for col_num, col_name in enumerate(df.columns, 1):
                    if 'Signal' in str(col_name):
                        signal_col = col_num
                        break
                
                if signal_col:
                    # Apply conditional formatting
                    worksheet.conditional_formatting.add(
                        f'{chr(64 + signal_col)}2:{chr(64 + signal_col)}{len(df) + 1}',
                        CellIsRule(operator='equal', formula=['"BUY"'], fill=buy_fill)
                    )
                    worksheet.conditional_formatting.add(
                        f'{chr(64 + signal_col)}2:{chr(64 + signal_col)}{len(df) + 1}',
                        CellIsRule(operator='equal', formula=['"HOLD"'], fill=hold_fill)
                    )
                    worksheet.conditional_formatting.add(
                        f'{chr(64 + signal_col)}2:{chr(64 + signal_col)}{len(df) + 1}',
                        CellIsRule(operator='equal', formula=['"SELL"'], fill=sell_fill)
                    )
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error exporting to Excel: {str(e)}")
            return b""


class SummaryStats:
    """Handles summary statistics creation and display"""
    
    @staticmethod
    def create_summary(signal_counts: Dict, total_tickers: int, avg_score: float) -> Dict:
        """
        Create summary statistics
        
        Args:
            signal_counts: Dictionary with signal counts
            total_tickers: Total number of tickers analyzed
            avg_score: Average weighted score
            
        Returns:
            Summary statistics dictionary
        """
        try:
            summary = {
                'total_tickers': total_tickers,
                'signal_distribution': signal_counts,
                'average_score': avg_score,
                'buy_percentage': (signal_counts.get('BUY', 0) / total_tickers * 100) if total_tickers > 0 else 0,
                'hold_percentage': (signal_counts.get('HOLD', 0) / total_tickers * 100) if total_tickers > 0 else 0,
                'sell_percentage': (signal_counts.get('SELL', 0) / total_tickers * 100) if total_tickers > 0 else 0
            }
            return summary
            
        except Exception as e:
            st.error(f"Error creating summary: {str(e)}")
            return {}
    
    @staticmethod
    def display_summary(summary: Dict):
        """
        Display enhanced summary statistics in Streamlit with beautiful design
        
        Args:
            summary: Summary statistics dictionary
        """
        try:
            if not summary:
                return
            
            # Enhanced header with custom styling
            st.markdown("""
            <div style="
                background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
                padding: 1rem;
                border-radius: 10px;
                margin-bottom: 1.5rem;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            ">
                <h2 style="
                    color: white;
                    margin: 0;
                    text-align: center;
                    font-weight: 600;
                    font-size: 1.5rem;
                ">📊 Summary Statistics</h2>
            </div>
            """, unsafe_allow_html=True)
            
            # Get values
            total_tickers = summary.get('total_tickers', 0)
            successful = summary.get('signal_distribution', {})
            errors = total_tickers - sum(successful.values()) if total_tickers > 0 else 0
            avg_score = summary.get('average_score', 0)
            
            # Calculate score range
            if 'score_range' in summary:
                score_range = summary['score_range']
            else:
                score_range = "N/A"
            
            # Top metrics row with enhanced styling
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    padding: 1.5rem;
                    border-radius: 15px;
                    text-align: center;
                    color: white;
                    box-shadow: 0 8px 16px rgba(102, 126, 234, 0.3);
                    margin-bottom: 1rem;
                ">
                    <h3 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">Total Analyzed</h3>
                    <h1 style="margin: 0.5rem 0 0 0; font-size: 2.5rem; font-weight: 700;">{}</h1>
                </div>
                """.format(total_tickers), unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
                    padding: 1.5rem;
                    border-radius: 15px;
                    text-align: center;
                    color: white;
                    box-shadow: 0 8px 16px rgba(17, 153, 142, 0.3);
                    margin-bottom: 1rem;
                ">
                    <h3 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">Successful</h3>
                    <h1 style="margin: 0.5rem 0 0 0; font-size: 2.5rem; font-weight: 700;">{}</h1>
                </div>
                """.format(sum(successful.values())), unsafe_allow_html=True)
            
            with col3:
                error_color = "#e74c3c" if errors > 0 else "#95a5a6"
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, {} 0%, {} 100%);
                    padding: 1.5rem;
                    border-radius: 15px;
                    text-align: center;
                    color: white;
                    box-shadow: 0 8px 16px rgba(231, 76, 60, 0.3);
                    margin-bottom: 1rem;
                ">
                    <h3 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">Errors</h3>
                    <h1 style="margin: 0.5rem 0 0 0; font-size: 2.5rem; font-weight: 700;">{}</h1>
                </div>
                """.format(error_color, error_color, errors), unsafe_allow_html=True)
            
            # Signal Distribution with enhanced cards
            st.markdown("""
            <div style="
                background: #f8f9fa;
                padding: 1.5rem;
                border-radius: 15px;
                margin: 1.5rem 0;
                border: 1px solid #e9ecef;
            ">
                <h3 style="
                    color: #495057;
                    margin: 0 0 1rem 0;
                    font-weight: 600;
                    text-align: center;
                ">Signal Distribution:</h3>
            """, unsafe_allow_html=True)
            
            # Signal cards
            col1, col2, col3 = st.columns(3)
            
            buy_count = successful.get('BUY', 0)
            hold_count = successful.get('HOLD', 0)
            sell_count = successful.get('SELL', 0)
            
            with col1:
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
                    padding: 1.5rem;
                    border-radius: 12px;
                    text-align: center;
                    color: white;
                    box-shadow: 0 6px 12px rgba(46, 204, 113, 0.3);
                ">
                    <h4 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">BUY</h4>
                    <h2 style="margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: 700;">{}</h2>
                </div>
                """.format(buy_count), unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #f39c12 0%, #e67e22 100%);
                    padding: 1.5rem;
                    border-radius: 12px;
                    text-align: center;
                    color: white;
                    box-shadow: 0 6px 12px rgba(243, 156, 18, 0.3);
                ">
                    <h4 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">HOLD</h4>
                    <h2 style="margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: 700;">{}</h2>
                </div>
                """.format(hold_count), unsafe_allow_html=True)
            
            with col3:
                st.markdown("""
                <div style="
                    background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
                    padding: 1.5rem;
                    border-radius: 12px;
                    text-align: center;
                    color: white;
                    box-shadow: 0 6px 12px rgba(231, 76, 60, 0.3);
                ">
                    <h4 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">SELL</h4>
                    <h2 style="margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: 700;">{}</h2>
                </div>
                """.format(sell_count), unsafe_allow_html=True)
            
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Score Statistics with enhanced styling
            st.markdown("""
            <div style="
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 1.5rem;
                border-radius: 15px;
                margin: 1.5rem 0;
                color: white;
                box-shadow: 0 8px 16px rgba(102, 126, 234, 0.3);
            ">
                <h3 style="
                    margin: 0 0 1rem 0;
                    font-weight: 600;
                    text-align: center;
                ">Score Statistics:</h3>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("""
                <div style="text-align: center; padding: 1rem;">
                    <h4 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">Average Score</h4>
                    <h2 style="margin: 0.5rem 0 0 0; font-size: 1.8rem; font-weight: 700;">{:.4f}</h2>
                </div>
                """.format(avg_score), unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div style="text-align: center; padding: 1rem;">
                    <h4 style="margin: 0; font-size: 0.9rem; opacity: 0.9;">Score Range</h4>
                    <h2 style="margin: 0.5rem 0 0 0; font-size: 1.2rem; font-weight: 700;">{}</h2>
                </div>
                """.format(score_range), unsafe_allow_html=True)
            
            st.markdown("</div>", unsafe_allow_html=True)
            
            # Enhanced signal distribution chart
            if summary.get('signal_distribution'):
                signal_data = summary['signal_distribution']
                if any(signal_data.values()):
                    st.markdown("""
                    <div style="
                        background: white;
                        padding: 1.5rem;
                        border-radius: 15px;
                        margin: 1.5rem 0;
                        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                        border: 1px solid #e9ecef;
                    ">
                        <h3 style="
                            color: #495057;
                            margin: 0 0 1rem 0;
                            font-weight: 600;
                            text-align: center;
                        ">📈 Signal Distribution Chart</h3>
                    """, unsafe_allow_html=True)
                    
                    # Create a more colorful chart
                    import pandas as pd
                    chart_df = pd.DataFrame(list(signal_data.items()), columns=['Signal', 'Count'])
                    st.bar_chart(chart_df.set_index('Signal'), use_container_width=True)
                    
                    st.markdown("</div>", unsafe_allow_html=True)
                    
        except Exception as e:
            st.error(f"Error displaying summary: {str(e)}")
